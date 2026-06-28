"""
EverBatt Excel Model Reader
===========================
修复 xlsm 文件中的 XML 兼容性问题，然后用 openpyxl 读取所有工作表内容。
"""

import zipfile, os, shutil, tempfile, glob
from xml.etree import ElementTree as ET


def fix_xlsm(src_path):
    """修复 EverBatt xlsm 文件中的 XML 非法值，返回修复后的临时文件路径"""
    base = os.path.join(tempfile.gettempdir(), 'everbatt_extract')
    if os.path.exists(base):
        shutil.rmtree(base)

    # 1. 解压
    print(f"解压: {src_path}")
    with zipfile.ZipFile(src_path, 'r') as zin:
        zin.extractall(base)

    # 2. 修复所有 XML
    fixed_count = 0
    for dirpath, dirnames, filenames in os.walk(base):
        for f in filenames:
            if not f.endswith(('.xml', '.rels')):
                continue
            full = os.path.join(dirpath, f)
            try:
                tree = ET.parse(full)
                root = tree.getroot()
                modified = False

                for el in root.iter():
                    # font family val > 14 → 改成 2
                    if el.tag.endswith('}family') or el.tag == 'family':
                        val = el.get('val', '0')
                        try:
                            if int(val) > 14:
                                el.set('val', '2')
                                modified = True
                        except ValueError:
                            pass
                    # theme > 10 → 改成 0
                    if 'theme' in el.attrib:
                        try:
                            if int(el.get('theme', '0')) > 10:
                                el.set('theme', '0')
                                modified = True
                        except ValueError:
                            pass
                    # indexed color > 65 → 改成 0
                    if 'indexed' in el.attrib:
                        try:
                            if int(el.get('indexed', '0')) > 65:
                                el.set('indexed', '0')
                                modified = True
                        except ValueError:
                            pass

                if modified:
                    tree.write(full, xml_declaration=True, encoding='UTF-8')
                    fixed_count += 1
            except Exception:
                pass

    print(f"修了 {fixed_count} 个 XML 文件")

    # 3. 重新打包
    tmp = os.path.join(tempfile.gettempdir(), 'everbatt_fixed.xlsm')
    if os.path.exists(tmp):
        os.remove(tmp)
    with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
        for dirpath, dirnames, filenames in os.walk(base):
            for f in filenames:
                full = os.path.join(dirpath, f)
                arcname = os.path.relpath(full, base)
                zout.write(full, arcname)

    print(f"修复后文件: {tmp} ({os.path.getsize(tmp):,} bytes)")
    return tmp


def read_all_sheets(filepath, max_rows=30, max_cols=8):
    """读取并打印所有工作表的前若干行"""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    print(f"\n共 {len(wb.sheetnames)} 个工作表:\n")

    for name in wb.sheetnames:
        ws = wb[name]
        print(f"{'='*60}")
        print(f"  {name}  (rows={ws.max_row}, cols={ws.max_column})")
        print(f"{'='*60}")
        for row in ws.iter_rows(min_row=1, max_row=min(max_rows, ws.max_row),
                                max_col=max_cols, values_only=False):
            vals = [str(c.value)[:100] if c.value is not None else '' for c in row]
            line = ' | '.join(vals)
            if line.strip(' |'):
                print(f"  R{row[0].row}: {line}")
        if ws.max_row > max_rows:
            print(f"  ... (还有 {ws.max_row - max_rows} 行，已省略)")
        print()
    wb.close()


def find_value(filepath, keyword):
    """在全部工作表中搜索包含关键词的单元格"""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    matches = []
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                max_col=ws.max_column, values_only=False):
            for cell in row:
                if cell.value and keyword.lower() in str(cell.value).lower():
                    matches.append((name, cell.row, cell.column_letter, cell.value))
    wb.close()
    for sheet, row, col, val in matches[:50]:
        print(f"  [{sheet}] R{row}C{col}: {str(val)[:120]}")
    print(f"\n共找到 {len(matches)} 个匹配")


if __name__ == '__main__':
    import sys

    # 默认读取 2023 版
    src = 'EverBatt 2023.xlsm'

    if len(sys.argv) > 1:
        src = sys.argv[1]

    # Step 1: 修复
    fixed = fix_xlsm(src)

    # Step 2: 读取全部工作表（前 30 行）
    print("\n" + "=" * 60)
    print("  读取所有工作表内容")
    print("=" * 60)
    read_all_sheets(fixed, max_rows=30, max_cols=8)

    # Step 3: 搜索关键数据
    print("\n" + "=" * 60)
    print("  搜索 'LFP' 相关数据")
    print("=" * 60)
    find_value(fixed, 'LFP')
